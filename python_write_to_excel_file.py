import requests
import pandas as pd
import time
import schedule
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

url = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'
headers = {
    'X-CMC_PRO_API_KEY': '42cbb389-b6a1-4fa5-9fe4-47b72054fcb8'
}
excel_file = "crypto_data.xlsx"

def fetch_crypto_data():
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()
    crypto_list = []
    for crypto in data['data'][:50]:
        crypto_list.append({
            "Name": crypto["name"],
            "Symbol": crypto["symbol"],
            "Price (USD)": crypto["quote"]["USD"]["price"],
            "Market Cap (USD)": crypto["quote"]["USD"]["market_cap"],
            "24h Volume (USD)": crypto["quote"]["USD"]["volume_24h"],
            "24h Change (%)": crypto["quote"]["USD"]["percent_change_24h"]
        })
    return pd.DataFrame(crypto_list)

def analyze_crypto_data(df):
    top_5_by_market_cap = df.nlargest(5, 'Market Cap (USD)')
    avg_price = df['Price (USD)'].mean()
    highest_change = df.nlargest(1, '24h Change (%)')
    lowest_change = df.nsmallest(1, '24h Change (%)')
    analysis_data = {
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Top 5 by Market Cap": ", ".join(top_5_by_market_cap['Name'].tolist()),
        "Average Price (USD)": avg_price,
        "Highest 24h Change": f"{highest_change.iloc[0]['Name']} ({highest_change.iloc[0]['24h Change (%)']}%)",
        "Lowest 24h Change": f"{lowest_change.iloc[0]['Name']} ({lowest_change.iloc[0]['24h Change (%)']}%)"
    }
    return analysis_data

def write_to_excel(dataframe, analysis_data):
    pd.set_option('display.float_format', lambda x: '%.8f' % x)
    pd.set_option('display.max_columns', None)
    
    dataframe['Market Cap (USD)'] = dataframe['Market Cap (USD)'].astype('int64')
    dataframe['24h Volume (USD)'] = dataframe['24h Volume (USD)'].astype('int64')
    dataframe['Price (USD)'] = dataframe['Price (USD)'].apply(lambda x: '{:.8f}'.format(x))
    dataframe['24h Change (%)'] = dataframe['24h Change (%)'].apply(lambda x: '{:.2f}'.format(x))

    if not os.path.exists(excel_file):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Live Data"
        wb.create_sheet("Analysis")
        wb.save(excel_file)

    wb = load_workbook(excel_file)
    if "Live Data" not in wb.sheetnames:
        wb.create_sheet("Live Data")
    live_data_sheet = wb["Live Data"]
    live_data_sheet.delete_rows(1, live_data_sheet.max_row)
    live_data_sheet.append(list(dataframe.columns))
    for row in dataframe.itertuples(index=False):
        live_data_sheet.append(row)

    if "Analysis" not in wb.sheetnames:
        wb.create_sheet("Analysis")
    analysis_sheet = wb["Analysis"]
    analysis_sheet.delete_rows(1, analysis_sheet.max_row)
    analysis_sheet.append(["Metric", "Value"])
    for key, value in analysis_data.items():
        analysis_sheet.append([key, value])
    
    wb.save(excel_file)

def update_excel():
    try:
        crypto_df = fetch_crypto_data()
        analysis_data = analyze_crypto_data(crypto_df)
        write_to_excel(crypto_df, analysis_data)
        print(f"Data updated at {analysis_data['Timestamp']}")
    except Exception as e:
        print(f"Error updating Excel: {str(e)}")

def main():
    update_excel()
    schedule.every(300).seconds.do(update_excel)
    print("Starting cryptocurrency tracker. Press Ctrl+C to stop.")
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == "__main__":
    main()
